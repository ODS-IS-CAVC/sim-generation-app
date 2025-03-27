import argparse
import json
import os

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as draw_image
from openpyxl.styles import Alignment, Border, Side, colors
from openpyxl.worksheet.worksheet import Worksheet

from commons import image_util
from commons.constants import *

# settings
WRK_DIR = os.path.abspath(os.path.dirname(__file__))
TEMPLATE_FILE = os.path.join(WRK_DIR, "input/template.xlsx")
OUTPUT_EXCEL_FILE_NAME = "summary_detection_distance.xlsx"


# index of the first line to insert in the excel file
EXCEL_DATA_FIRST_LINE_IDX = 3
HEADER_COLS = [
    "画像",
    "画像ファイル名",
    "水平方向視野角(θ°)",
    "OBJ_ID",
    "画像の解像度(W)",
    "画像の解像度(H)",
    "垂直方向の視野角(Φ°)",
    "検出した車の矩形左下部の座標 (x)",
    "検出した車の矩形左下部の座標 (y)",
    "検出した車の矩形左下部までの水平方向角度(ψx)",
    "検出した車の矩形左下部までの視点の角度 (ψ°)",
    "車までのx方向の距離(dx(m))",
    "車までの奥行き方向距離(d (m))",
    "車までの距離(dx^2+d^2)^1/2",
]


# function to add data to the new template
def add_data(
    row_idx: int, ws: Worksheet, data: list, first_row_data: bool = False
):
    # data: image_fpath, image_fn, theta, OBJ_ID, w, h, phi_deg, x, y, psi_x_deg, psi_deg, dx, dy, distance_to_car
    ws[f"C{row_idx}"].value = data[1]
    ws[f"D{row_idx}"].value = data[2]
    ws[f"E{row_idx}"].value = data[3]
    ws[f"F{row_idx}"].value = data[4]
    ws[f"G{row_idx}"].value = data[5]
    ws[f"H{row_idx}"].value = data[6]
    ws[f"I{row_idx}"].value = data[7]
    ws[f"J{row_idx}"].value = data[8]
    ws[f"K{row_idx}"].value = data[9]
    ws[f"L{row_idx}"].value = data[10]
    ws[f"M{row_idx}"].value = data[11]
    ws[f"N{row_idx}"].value = data[12]
    ws[f"O{row_idx}"].value = data[13]

    input_image_path = data[0]
    image = image_util.read_image(input_image_path)
    h, w, c = image.shape
    image_ratio = w / h

    # set col & row size
    COL_W = 50  # 50cm
    ROW_H = COL_W / image_ratio
    COL_OFFSET = 5
    ROW_OFFSET = 10
    ws.row_dimensions[row_idx].height = ROW_H * 6 + ROW_OFFSET
    ws.column_dimensions["B"].width = COL_W + COL_OFFSET

    # set image size
    IMG_W = 400  # pixel
    IMG_H = IMG_W / image_ratio

    # insert image
    img = draw_image(data[0])
    img.height = IMG_H
    img.width = IMG_W

    # 画像を貼り付ける
    ws.add_image(img, f"B{row_idx}")

    col_names = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
    for ii, col in enumerate(col_names):
        cell = ws.cell(row_idx, 2 + ii)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        border_side = Side("thin", color=colors.BLACK)
        border_top = (
            Side("medium", color=colors.BLACK) if first_row_data else border_side
        )
        cell.border = Border(
            left=border_side, right=border_side, top=border_top, bottom=border_side
        )


def json2dataframe(rel_coord_file, column_names):
    """Json to dataframe

    Args:
        rel_coord_file (str): 自車の相対座標と検出車の相対座標のjsonファイルパス（infer_distance_to_car.pyの出力json)
        column_names (list[string]): 列名リスト
    """
    # Read json file
    with open(rel_coord_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    camera_parameter = data["camera_parameter"]
    theta = camera_parameter["aov_horizontal"]
    phi_deg = camera_parameter["aov_vertical"]
    w, h = camera_parameter["image_size"][0], camera_parameter["image_size"][1]
    rel_coord_results = data["results"]

    # Loop over results
    output_data = []
    for item in rel_coord_results:
        output_path = item["file"]
        image_fn = os.path.basename(output_path)

        for obj in item["detections"]:
            detection_point = obj["detection_point"]
            x, y = detection_point[0], detection_point[1]
            psi_x_deg, psi_y_deg = obj["angle"]
            distance = obj["distance"]
            dx, dy, d = distance
            obj_id = obj["obj_id"]
            output_data.append(
                [
                    output_path,
                    image_fn,
                    theta,
                    obj_id,
                    w,
                    h,
                    phi_deg,
                    x,
                    y,
                    psi_x_deg,
                    psi_y_deg,
                    dx,
                    dy,
                    d,
                ]
            )

    df = pd.DataFrame(columns=column_names, data=output_data)
    return df


def save_excel_file(template_file, output_path, df_data, start_row):
    # opening the destination excel file
    wb = openpyxl.load_workbook(template_file)

    # wb._active_sheet_index = 1
    ws = wb.active

    # saving the destination excel file
    id_row = start_row
    for idx in range(0, df_data.shape[0], 1):
        data = df_data.values.tolist()[idx]
        add_data(id_row, ws, data, first_row_data=(idx == 0))
        id_row = id_row + 1

    # update sheet name
    theta = df_data["水平方向視野角(θ°)"][0]
    ws.title = ws.title.split("_")[0] + "_" + str(int(theta)) + "°"

    wb.save(str(output_path))


def main():
    # 引数をパースする
    parser = argparse.ArgumentParser(description="相対座標の推定結果をエクセルとしてまとめる")
    parser.add_argument(
        "--rel_coord_file",
        type=str,
        required=True,
        help="自車の相対座標と検出車の相対座標のjsonファイルパス（infer_distance_to_car.pyの出力json）",
    )
    parser.add_argument(
        "--output_dir",
        type=str,
        required=True,
        default="output",
        help="出力エクセルのフォルダーパス",
    )
    
    # parse input arguments
    args = parser.parse_args()

    rel_coord_file = args.rel_coord_file
    assert isinstance(rel_coord_file, str)

    output_dir = args.output_dir
    assert isinstance(output_dir, str)

    # make output dir
    os.makedirs(output_dir, exist_ok=True)

    # execute
    output_excel_path = os.path.join(output_dir, OUTPUT_EXCEL_FILE_NAME)
    df_data = json2dataframe(rel_coord_file, HEADER_COLS)
    save_excel_file(
        TEMPLATE_FILE, output_excel_path, df_data, EXCEL_DATA_FIRST_LINE_IDX
    )


if __name__ == "__main__":
    main()
